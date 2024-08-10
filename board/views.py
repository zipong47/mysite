import os
import re
from datetime import datetime
import json

from mysite import settings
from board.common import get_request
from .models import *
from .form import EnvReportForm,EditTestPlanForm,DisplayEditTestPlanForm
from openpyxl import load_workbook
import openpyxl
from openpyxl.utils import get_column_letter

from django.template import loader
from django.shortcuts import get_object_or_404, render,redirect
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect,JsonResponse
from django.contrib import messages
from django.template.loader import render_to_string
from django.core.exceptions import ObjectDoesNotExist
# ban csrf
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.paginator import Paginator

# 定义字符串列表
STATIONS_STRINGS = ["ICT","DFU","FCT","SOC-TEST","WIFI-BT-COND-B","WIFI-BT-COND"]


class project_detail:
    def __init__(self,project_name,config_subproject_name_dict):
        self.total_subproject_nums=0
        self.project_name=project_name
        self.config_subproject_name_dict=config_subproject_name_dict
        self.subproject_num_list=[]
        for each in config_subproject_name_dict:
            temp_num=0
            for i in config_subproject_name_dict[each]:
                temp_num=temp_num+1
                self.total_subproject_nums=self.total_subproject_nums+1
            self.subproject_num_list.append(temp_num)

class build_unit:
    def __init__(self,project_name,config,build_name,checkin_nums):
        self.project_name=project_name
        self.config=config
        self.build_name=build_name
        self.checkin_nums=checkin_nums
        self.station_yield_list=[]
        
def index(request):
    project_name_set=Board.objects.values_list('project_name',flat=True).distinct()
    index_detail=[]
    for project_name in project_name_set:
        temp_dit={}
        config_set=Board.objects.filter(project_name=project_name).values_list('project_config',flat=True).distinct()
        for each in config_set:
            temp_list=[]
            build_name_set=Board.objects.filter(project_name=project_name,project_config=each).values_list('subproject_name',flat=True).distinct()
            for build_name in build_name_set:
                temp_list.append(build_name)
            temp_dit[each]=temp_list
        b=project_detail(project_name=project_name,config_subproject_name_dict=temp_dit)
        #index_detail.append(b)
    for project_name in project_name_set:
        config_set=Board.objects.filter(project_name=project_name).values_list('project_config',flat=True).distinct()
        for config in config_set:
            build_name_set=Board.objects.filter(project_name=project_name,project_config=config).values_list('subproject_name',flat=True).distinct()
            for build_name in build_name_set:
                checkin_set=Board.objects.filter(project_name=project_name,project_config=config,subproject_name=build_name).filter(env_finished_flag=True)
                nums=len(checkin_set)
                bu=build_unit(project_name=project_name,config=config,build_name=build_name,checkin_nums=nums)
                station_type_list=["ICT","DFU","FCT","SOC-TEST","WIFI-BT-COND","WIFI-BT-COND-B","W3","SWDL"]
                for station_type in station_type_list:
                    station_num=0
                    for board_unit in checkin_set:
                        checkin_record=TestRecord.objects.filter(board=board_unit).filter(station_type='checkin').order_by("-start_time").first()
                        station_record=TestRecord.objects.filter(board=board_unit).filter(station_type=station_type).order_by("-start_time").first()
                        if checkin_record!=None:
                            print(checkin_record.start_time)
                        if station_record!=None:
                            print(station_record.start_time)
                        if(checkin_record != None and station_record != None and checkin_record.start_time < station_record.start_time and station_record.result == 'pass'):
                            station_num=station_num+1
                    bu.station_yield_list.append(station_num)
                    
                    # rate=(station_num/len(checkin_set))*100
                    # rate="{:.2f}%".format(rate)
                    # bu.station_yield_list.append(rate)
                index_detail.append(bu)
                print("nums*******:"+str(nums))
                
    context = {"details": index_detail,"num":5}
    return render(request, "board/index.html", context)

def smt(request):
    latest_question_list = ['physics', 'chemistry', 1997, 2000]
    template = loader.get_template("board/smt.html")
    context = {"latest_question_list": latest_question_list}
    return render(request, "board/smt.html", context)

def upload(request):
    print("upload被调用-------------------")
    return render(request, "board/upload.html")

def check_test_plan_status(test_records, test_plan):
    plan_index = 0
    extra_tests = []
    missing_tests = []
    sequence_errors = []
    current_tests = []
    
    for record in test_records:
        if plan_index < len(test_plan) and record.station_type == test_plan[plan_index]:
            plan_index += 1
            current_tests.append(record.station_type)
        elif plan_index < len(test_plan) and record.station_type != test_plan[plan_index]:
            sequence_errors.append(record.station_type)
        else:
            extra_tests.append(record.station_type)
        
    # Check for any remaining test plans that were not executed
    if plan_index < len(test_plan):
        missing_tests.extend(test_plan[plan_index:])

    return plan_index == len(test_plan), extra_tests, missing_tests, sequence_errors, current_tests
@csrf_exempt
def show_history(request,type):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':  # Ajax请求
        today=datetime.now().date()
        start_of_day = datetime.combine(today, datetime.min.time())
        end_of_day = datetime.combine(today, datetime.max.time())
        test_records_with_schedule=[]
        
        if type=='checkin':
            test_records_today = TestRecord.objects.filter(station_type='checkin').filter(start_time__range=(start_of_day, end_of_day)).order_by('-start_time')
        elif type=='checkout':
            test_records_today = TestRecord.objects.filter(station_type='checkout').filter(start_time__range=(start_of_day, end_of_day)).order_by('-start_time')
            
        for record in test_records_today:
            try:
                schedule = TestSchedule.objects.get(serial_number=record.board,cp_nums=record.cp_nums)
                next_station = schedule.test_sequence.split('→')[0]
            except TestSchedule.DoesNotExist:
                schedule = None
            test_records_with_schedule.append((record, schedule, next_station))

        daily_table_context={}
        daily_table_context["test_records_with_schedule"]=test_records_with_schedule
        if type=='checkin':
            daily_table_html=render_to_string('board/checkin_daily_table.html',daily_table_context,request=request)
            return JsonResponse({"daily_table_html":daily_table_html})
        else:
            daily_table_html=render_to_string('board/checkout_daily_table.html',daily_table_context,request=request)
            return JsonResponse({"daily_table_html":daily_table_html})

    messages.error(request,"服务器错误！")
    if type=='checkin':
        return JsonResponse({'redirect_url': 'board/checkin.html/'}, status=400)
    else:
        return JsonResponse({'redirect_url': 'board/checkout.html/'}, status=400)

@csrf_exempt
def check_in(request):
    context={}
    return render(request, "board/checkin.html",context)

@csrf_exempt
def checkin_ajax(request,sn_str):
    context={}
    print("____checkin_ajax____")
    message=""
    try:
        board=Board.objects.get(pk=sn_str)
    except Board.DoesNotExist:
        messages.error(request,"未在数据库中查到关于该MLB的相关信息! 请检查sn是否正确。\n")
        return JsonResponse({'redirect_url': '/board/checkin'}, status=400)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':  # Ajax请求
        current_cp=board.cp_nums
        try:
            test_schedule=TestSchedule.objects.filter(serial_number=board,cp_nums=current_cp).get()
        except TestSchedule.DoesNotExist:
            messages.error(request,"未在数据库中查到关于该MLB的测试计划!\n")
            return JsonResponse({'redirect_url': '/board/checkin'}, status=400)

        test_plan_list=test_schedule.test_sequence.split("→")
        test_plan_list=['checkin']+test_plan_list
        
        if_checkin = TestRecord.objects.filter(board=board,cp_nums=current_cp,station_type='checkin').exists()
        if if_checkin:
            message=message+"该主板已经Check-In了，请不要重复扫描。\n"
        else:
            temp_record=TestRecord(board=board,station_type="checkin",cp_nums=board.cp_nums,start_time=datetime.now(),stop_time=datetime.now(),result='pass')
            temp_record.save()
            board.env_finished_flag=True
            board.save()

        test_records = TestRecord.objects.filter(board=board,cp_nums=current_cp,result='pass').order_by('start_time')
        is_followed, extra_tests, missing_tests, sequence_errors, current_tests = check_test_plan_status(test_records, test_plan_list)

        test_plan_order_list=[]
        if sequence_errors:
            print("漏测或者测试顺序错误: ", sequence_errors)
            test_plan_order_list=current_tests+sequence_errors
            print(test_plan_order_list)
            message=message+"该主板测试异常，存在漏测或者测试顺序错误，请检查。\n "
        elif missing_tests:
            print("未测工站: ", missing_tests)
            test_plan_order_list=current_tests
            message=message+f"主板NO.{board.board_number}在CP{current_cp}下一个工站：{missing_tests[0]}。\n "
        elif extra_tests:
            print("多测工站: ", extra_tests)
            test_plan_order_list=current_tests
            message=message+"该主板多测了工站，请检查是否正常。\n "
        elif is_followed:
            print("所有测试通过，无错误")
            message=message+f"所有测试通过，无错误。主板在CP{current_cp}的测试都已经完成！\n "
            test_plan_order_list=current_tests+['checkout']
        
        test_plan_list=test_plan_list+['checkout']

        test_history=TestRecord.objects.filter(board=board,cp_nums=current_cp).order_by('start_time')
        if test_history.exists():
            context["test_history"]=test_history
        
        context["message"]=message
        context["test_plan"]=test_plan_list
        context["board_info"]=board

        checkin_table_html=render_to_string('board/checkin_table.html',context,request=request)

        # ------------- daily table ---------------
        today=datetime.now().date()
        start_of_day = datetime.combine(today, datetime.min.time())
        end_of_day = datetime.combine(today, datetime.max.time())
        test_records_with_schedule=[]
        test_records_today = TestRecord.objects.filter(station_type='checkin').filter(start_time__range=(start_of_day, end_of_day)).order_by('-start_time')
        for record in test_records_today:
            try:
                schedule = TestSchedule.objects.get(serial_number=record.board,cp_nums=record.cp_nums)
                next_station = schedule.test_sequence.split('→')[0]
            except TestSchedule.DoesNotExist:
                schedule = None
            test_records_with_schedule.append((record, schedule, next_station))

        daily_table_context={}
        daily_table_context["test_records_with_schedule"]=test_records_with_schedule
        daily_table_html=render_to_string('board/checkin_daily_table.html',daily_table_context,request=request)
        
        print("before return~~~~~~~~~~~")
        
        return JsonResponse({"table":checkin_table_html,"test_plan_order_list":test_plan_order_list,"test_plan_list":test_plan_list,"daily_table_html":daily_table_html})

    messages.error(request,"服务器错误！")
    return JsonResponse({'redirect_url': 'board/checkin.html/'}, status=400)

@csrf_exempt
def check_out(request):
    context={}
    return render(request, "board/checkout.html",context)

@csrf_exempt
def checkout_ajax(request,sn_str):
    context={}
    print("____checkout_ajax____")
    message=""
    try:
        board=Board.objects.get(pk=sn_str)
    except Board.DoesNotExist:
        print("eception__________________")
        messages.error(request,"未在数据库中查到关于该MLB的相关信息! 请检查sn是否正确。\n")
        return JsonResponse({'redirect_url': '/board/checkout'}, status=400)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':  # Ajax请求
        current_cp=board.cp_nums
        try:
            test_schedule=TestSchedule.objects.filter(serial_number=board,cp_nums=current_cp).get()
        except TestSchedule.DoesNotExist:
            messages.error(request,"未在数据库中查到关于该MLB的test plan相关信息!\n")
            return JsonResponse({'redirect_url': 'board/checkout.html/'}, status=400)

        test_plan_list=test_schedule.test_sequence.split("→")
        test_plan_list=['checkin']+test_plan_list

        test_records = TestRecord.objects.filter(board=board,cp_nums=current_cp,result__in=['pass','cof']).order_by('start_time')
        is_followed, extra_tests, missing_tests, sequence_errors, current_tests= check_test_plan_status(test_records, test_plan_list)
        
        test_plan_order_list=[]
        if sequence_errors:
            print("漏测或者测试顺序错误: ", sequence_errors)
            test_plan_order_list=current_tests+sequence_errors
            print(test_plan_order_list)
            message=message+"该主板测试异常，存在漏测或者测试顺序错误，请检查。\n "
        elif missing_tests:
            print("未测工站: ", missing_tests)
            test_plan_order_list=current_tests
            message=message+f"该主板在CP{current_cp}还有未完成的测试，请检查。\n "
        elif extra_tests:
            print("多测工站: ", extra_tests)
            test_plan_order_list=current_tests
            message=message+"该主板多测了工站，请检查是否正常。\n "
        elif is_followed:
            print("所有测试通过，无错误")
            message=message+f"所有测试通过，无错误。主板在CP{current_cp}checkout成功！\n "
            if_test_pass=True
            test_plan_order_list=current_tests+['checkout']
            checkin_record_set=TestRecord.objects.filter(board=board).filter(station_type="checkin")
            checkout_record_set=TestRecord.objects.filter(board=board).filter(station_type="checkout")
            if_checkout = TestRecord.objects.filter(board=board,cp_nums=current_cp,station_type='checkout').exists()
            if(len(checkin_record_set)>len(checkout_record_set) and not if_checkout):
                temp_record=TestRecord(board=board,station_type="checkout",cp_nums=board.cp_nums,start_time=datetime.now(),stop_time=datetime.now(),result='pass')
                temp_record.save()
                board.cp_nums=board.cp_nums+100
                board.env_finished_flag=False
                board.if_overdue=False
                board.save()
            else:
                message=message+"该板子已经checkout了! \n"
                context["message"]=message
        
        test_plan_list=test_plan_list+['checkout']

        test_history=TestRecord.objects.filter(board=board,cp_nums=current_cp).order_by('start_time')
        if test_history.exists():
            context["test_history"]=test_history

        context["message"]=message
        context["test_plan"]=test_plan_list
        context["board_info"]=board
        
        checkin_table_html=render_to_string('board/checkin_table.html',context,request=request)

        # ------------- daily table ---------------
        today=datetime.now().date()
        start_of_day = datetime.combine(today, datetime.min.time())
        end_of_day = datetime.combine(today, datetime.max.time())
        test_records_with_schedule=[]
        test_records_today = TestRecord.objects.filter(station_type='checkout').filter(start_time__range=(start_of_day, end_of_day)).order_by('-start_time')
        for record in test_records_today:
            try:
                schedule = TestSchedule.objects.get(serial_number=record.board,cp_nums=record.cp_nums)
                print(f"current cp:{record.cp_nums}")
                next_station = schedule.test_sequence.split('→')[0]
                
                
            except TestSchedule.DoesNotExist:
                schedule = None
            test_records_with_schedule.append((record, schedule, next_station))

        daily_table_context={}
        daily_table_context["test_records_with_schedule"]=test_records_with_schedule
        daily_table_html=render_to_string('board/checkout_daily_table.html',daily_table_context,request=request)
        
        print("before return~~~~~~~~~~~")
        
        return JsonResponse({"table":checkin_table_html,"test_plan_order_list":test_plan_order_list,"test_plan_list":test_plan_list,"daily_table_html":daily_table_html})

    messages.error(request,"服务器错误！")
    return JsonResponse({'redirect_url': 'board/checkout.html/'}, status=400)
  

def handle_cp_str(cp_str):
    str_list=cp_str.split(",")
    numbers = re.findall(r'\d+', str_list[-1])
    return numbers[0]

@csrf_exempt
def get_env_report(request):
    form = EnvReportForm()
    if request.method == "POST":
        form = EnvReportForm(request.POST,request.FILES)
        if form.is_valid():
            # info_dic=form.cleaned_data['project_name_select_field']
            project_name=request.POST['project_name_select_field']
            project_config=request.POST['project_config']
            subproject_name=request.POST['subproject_field']
            file_data=request.FILES['env_report_file_field']
            file_absolute_path=os.path.join(settings.BASE_DIR, 'board\\static\\download\\')
            file_absolute_path=os.path.join(file_absolute_path, datetime.now().strftime("%Y%m%d_%H%M%S_")+file_data.name)
            print(file_absolute_path)
            with open(file_absolute_path, 'wb+') as destination:
                for chunk in file_data.chunks():
                    destination.write(chunk)
            messages.success(request,"上传成功！")
            workbook=load_workbook(file_absolute_path)
            worksheet=workbook['sheet1']
            total_rows=worksheet.max_row
            exist_board_list=[]
            # [7] APN、[8] HHPN、[9] First GS SN、[10] Second GS SN、[11] Site、[12] Product Code
            for each in worksheet.iter_rows(min_row=1):
                board_num=each[2].value
                configuration=each[3].value
                sn=each[4].value
                test_item=each[5].value
                cp_num=handle_cp_str(each[6].value)
                apn=each[7].value
                hhpn=each[8].value
                first_gs_sn=each[9].value
                second_gs_sn=each[10].value
                site=each[11].value
                product_code=each[12].value
                if not Board.objects.filter(serial_number=sn).exists():
                    a_board=Board(project_name=project_name,project_config=project_config,subproject_name=subproject_name,
                            serial_number=sn,configuration=configuration,board_number=board_num,
                            test_item_name=test_item,cp_nums=cp_num,product_code=product_code,APN=apn,HHPN=hhpn,
                            first_GS_sn=first_gs_sn,second_GS_sn=second_gs_sn,site=site)
                    a_board.save()
                else:
                    exist_board_list.append(sn)
                        
            return HttpResponseRedirect(request.path, {"form": form, "exist_board_list": exist_board_list})
    else:
        form = EnvReportForm()
    return render(request, "board/upload_report.html", {"form": form})

@csrf_exempt
def update_summary(request):
    checkin_set=Board.objects.filter(env_finished_flag=True)
    IP="172.16.243.140:80"
    # station_type_list=["ICT","DFU","FCT","SOC-TEST","WIFI-BT-COND-B","WIFI-BT-COND"]
    station_type_list=["ICT"]
    print("*******update summary")
    for board_unit in checkin_set:
        # cp_num=board_unit.cp_nums
        # TestRecord.objects.filter(serial_number=board_unit.serial_number).filter(cp_nums=cp_num)
        for station_type in station_type_list:
            result_dict=get_request(ip=IP,sn=board_unit.serial_number,station_type=station_type)
            result_dict={}
            if board_unit.serial_number == 'sn7':
                result_dict={
                    'start_time':'2024-05-14 16:03:00',
                    'stop_time':'2024-05-14 17:03:00',
                    'result':'pass'}
            else:
                result_dict={'start_time':'','stop_time':'','result':''}
            
            lastest_record=TestRecord.objects.filter(board=board_unit).filter(station_type=station_type).order_by("-start_time").first()

            print("lastest_record:")
            print(lastest_record)
            normal_flag=True
            for key,value in result_dict.items():
                print("value:"+value)
                if value == "" or value == None:
                    normal_flag=False
                    break
            
            if normal_flag:
                starttime = datetime.strptime(result_dict["start_time"], "%Y-%m-%d %H:%M:%S")
                if (lastest_record != None and starttime <= lastest_record.start_time):
                    print("没记录")
                    continue
                # if lastest_record.result == 'fail':
                #     lastest_record.delete()
                #     print("已经删除失败的测试结果。")
                start_time = datetime.strptime(result_dict["start_time"], "%Y-%m-%d %H:%M:%S")
                stop_time = datetime.strptime(result_dict["stop_time"], "%Y-%m-%d %H:%M:%S")
                record=TestRecord(board=board_unit,station_type=station_type,start_time=start_time,stop_time=stop_time,result=result_dict["result"],cp_nums=board_unit.cp_nums)
                record.save()
    
    return redirect("index")

class test_plan_unit:
    def __init__(self,project_config_build,checkin_nums,board_no,test_plan):
        self.project_config_build=project_config_build
        self.checkin_nums=checkin_nums
        self.board_no=board_no
        self.test_plan=test_plan
        self.pk=""
    
    def __str__(self) -> str:
        return (self.project_config_build+"::"+self.checkin_nums+"::"+self.board_no+"::"+self.test_plan)
    
@csrf_exempt
def enter_test_plan(request):
    project_config_build_list=[]
    project_name_set=Board.objects.values_list('project_name',flat=True).distinct()
    for project_name in project_name_set:
        config_set=Board.objects.filter(project_name=project_name).values_list('project_config',flat=True).distinct()
        for each in config_set:
            build_name_set=Board.objects.filter(project_name=project_name,project_config=each).values_list('subproject_name',flat=True).distinct()
            for build_name in build_name_set:
                # str=str(project_name)+str(each)+str(build_name)
                str=project_name+"-"+each+"-"+build_name
                project_config_build_list.append(str)
    
    context={'options': project_config_build_list}
    data = request.POST.getlist('data[]')
    project_config_build_str = request.POST.get('extra_data')
    checkpoint_str = request.POST.get('cp')

    if checkpoint_str!="" and checkpoint_str!=None:
        checkpoint_list=checkpoint_str.split(",")
        test_sequence="→".join(data)
        project_name,config_name,build_name=project_config_build_str.split("-")
        
        exist_test_schedule=[]
        test_plan_list=[]
        board_query_set=Board.objects.filter(project_name=project_name,project_config=config_name,subproject_name=build_name)
        for board in board_query_set:
            for cp in checkpoint_list:
                if TestSchedule.objects.filter(serial_number=board,cp_nums=cp).exists():
                    exist_test_schedule.append(board.board_number)
                    exist_test_schedule.append(cp)
                    continue
                new_record=TestSchedule(serial_number=board,cp_nums=cp,test_sequence=test_sequence)
                new_record.save()
                plan_unit=test_plan_unit(project_config_build_str,cp,board.board_number,test_sequence)
                plan_unit.pk=new_record.pk
                test_plan_list.append(plan_unit)
        context["test_plan_list"]=test_plan_list
        if len(exist_test_schedule)==0:
            messages.success(request," 成功上传测试计划！")
        else:
            temp_list=[]
            for i in range(0, len(exist_test_schedule), 2):
                str=f"{exist_test_schedule[i]}-{exist_test_schedule[i+1]}"
                temp_list.append(str)
            str=",".join(temp_list)
            messages.error(request,"成功上传部分测试计划！以下的 板号-CP 重复上传了数据："+str+" 请保存该信息后手动编辑！")
    print(context) 
    context['enter_flag']='enter_flag'
    return render(request, "board/enter_test_plan.html",context)

@csrf_exempt
def search_test_plan(request):
    project_config_build_list=[]
    project_name_set=Board.objects.values_list('project_name',flat=True).distinct()
    for project_name in project_name_set:
        config_set=Board.objects.filter(project_name=project_name).values_list('project_config',flat=True).distinct()
        for each in config_set:
            build_name_set=Board.objects.filter(project_name=project_name,project_config=each).values_list('subproject_name',flat=True).distinct()
            for build_name in build_name_set:
                str=project_name+"-"+each+"-"+build_name
                project_config_build_list.append(str)
    
    context={'options': project_config_build_list}
    
    data=request.POST.get("keyword")
    search_option=request.POST.get("flexRadioDefault")
    test_plan_list=[]
    if search_option=="sn":
        data_list=data.split()
        for sn in data_list:
            print("----------"+sn)
            try:
                board=Board.objects.get(serial_number=sn)
            except Board.DoesNotExist:
                messages.error(request,"不存在该SN的板子!请检查该SN是否存在:"+sn)
                return render(request, "board/enter_test_plan.html",context)
            
            str=board.project_name+"-"+board.project_config+"-"+board.subproject_name
            test_schedule_set=TestSchedule.objects.filter(serial_number=board)
            if len(test_schedule_set)==0:
                messages.error(request,"该sn的板子没有测试计划!")
            else:
                for each in test_schedule_set:
                    plan_unit=test_plan_unit(str,each.cp_nums,board.board_number,each.test_sequence)
                    plan_unit.pk=each.pk
                    test_plan_list.append(plan_unit)
                    print(each.pk)
    elif search_option=="build_name":
        try:
            project_name,config_name,build_name=data.split("-")
        except ValueError:
            messages.error(request,"正在使用Project-Configuration-Build搜索，但输入的格式有误，请使用符号(-)进行分割每个名称。")
            return render(request, "board/enter_test_plan.html",context)
        board_set=Board.objects.filter(project_name=project_name,project_config=config_name,subproject_name=build_name)
        if len(board_set) == 0:
            messages.error(request,"该build不存在，请确认是否已经录入或者是否输入错误!")
        else:   
            for board in board_set:
                str=board.project_name+"-"+board.project_config+"-"+board.subproject_name
                test_schedule_set=TestSchedule.objects.filter(serial_number=board)
                for each in test_schedule_set:
                    plan_unit=test_plan_unit(str,each.cp_nums,board.board_number,each.test_sequence)
                    test_plan_list.append(plan_unit)
    context["test_plan_list"]=test_plan_list
    return render(request, "board/enter_test_plan.html",context)

@csrf_exempt
def edit_test_plan(request,item_id):
    context={}
    try:
        test_schedule=TestSchedule.objects.get(pk=item_id)
    except TestSchedule.DoesNotExist:
        messages.error(request,"未在数据库中查到test schedule相关主键信息!")
        return render(request, "board/enter_test_plan.html",context)
        
    try:
        board=test_schedule.serial_number
    except Board.DoesNotExist:
        messages.error(request,"未在数据库中查到test schedule相关board的主键信息!")
        return render(request, "board/enter_test_plan.html",context)

    print("editing test plan......")

    if request.method == 'POST':
        form = EditTestPlanForm(request.POST,instance=test_schedule)
        if form.is_valid():
            # 获取提交的字段值
            submitted_value = form.cleaned_data['test_sequence']
            valid_flag=True
            for each in submitted_value.split("→"):
                if each not in STATIONS_STRINGS:
                    valid_flag=False
            if valid_flag:
                form.save()
                messages.success(request,f"板号:{board.board_number} 已成功提交和修改 Check Point:{test_schedule.cp_nums} 的测试计划。")
            else:
                messages.error(request, '提交的测试计划格式有问题。请检查修改后重新提交！')
            
            # 补全消息框后再实现下方功能
            # context["test_plan_list"]=test_plan_list
            return render(request, "board/enter_test_plan.html",context)
    else:
        form = EditTestPlanForm(instance=test_schedule)
        board_info_form = DisplayEditTestPlanForm(instance=board)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':  # Ajax请求
            form_html = render_to_string('board/edit_form_snippet.html', {'form': form,'board_info_form':board_info_form}, request=request)
            return JsonResponse({'form': form_html})  # 返回表单的HTML
        
    return render(request, 'board/enter_test_plan.html', {'form': form, 'item_id': item_id})

@csrf_exempt
def eception_page(request):
    project_name_set=Board.objects.values_list('project_name',flat=True).distinct()
    station_type_set = TestRecord.objects.exclude(station_type__in=['checkin', 'checkout']).values_list('station_type', flat=True).distinct()
    context={"project_name_list":list(project_name_set),"station_type_list":list(station_type_set)}
    return render(request, "board/eception_page.html",context)

@csrf_exempt
def search_eception(request):
    if request.method == 'POST':
        start_time = request.POST.get('startTime')
        end_time = request.POST.get('EndTime')
        project_name = request.POST.get('project')
        station_type = request.POST.get('station')
        
        # Initialize query filters
        test_record_filters = Q()
        board_filters = ~Q(status='testing') & ~Q(status='archived')
        
        # Add filters based on user input
        if start_time and end_time:
            test_record_filters &= Q(start_time__gte=start_time, stop_time__lte=end_time)
        
        if project_name:
            board_filters &= Q(project_name=project_name)
        
        if station_type:
            test_record_filters &= Q(station_type=station_type)
        
        # Retrieve Boards based on filters
        boards = Board.objects.filter(board_filters)
        for board in boards:
            print(f"board:{board.serial_number}")
        
        # Retrieve TestRecords based on filters and filter by the retrieved boards
        test_records = TestRecord.objects.filter(test_record_filters, board__in=boards)
        
        # Combine the results
        results = []
        for test_record in test_records:
            error_records = test_record.error_records.all()
            if error_records.exists():
                board = test_record.board
                for error_record in error_records:
                    results.append({
                        'board': {
                            'project_name': board.project_name,
                            'project_config': board.project_config,
                            'subproject_name': board.subproject_name,
                            'serial_number': board.serial_number,
                            'configuration': board.configuration,
                            'board_number': board.board_number,
                            'test_item_name': board.test_item_name,
                            'cp_nums': board.cp_nums,
                            'product_code': board.product_code,
                            'APN': board.APN,
                            'HHPN': board.HHPN,
                            'first_GS_sn': board.first_GS_sn,
                            'second_GS_sn': board.second_GS_sn,
                            'env_finished_flag': board.env_finished_flag,
                            'status': board.status,
                        },
                        'test_record': {
                            'station_type': test_record.station_type,
                            'start_time': test_record.start_time,
                            'cp_nums': test_record.cp_nums,
                            'stop_time': test_record.stop_time,
                            'result': test_record.result,
                            'operator': test_record.operator,
                            'remark': test_record.remark,
                        },
                        'error_record': {
                            'id': error_record.id,
                            'fail_message': error_record.fail_message,
                            'remark': error_record.remark,
                        }
                    })
                    print(f"error_record.id:{error_record.id}")
        
        if not results:
            print("results结果如下")
            print(results)
            customize_message_html = render_to_string('board/customize_message.html', {'customize_message': '未找到符合条件的记录'}, request=request)
            return JsonResponse({'results': results, 'customize_message': customize_message_html})

        return JsonResponse({'results': results})
    
    return JsonResponse({'error': 'Invalid request method.'}, status=400)

@csrf_exempt
def search_serial_number_in_eception_page(request):
    if request.method == 'POST':
        serial_number = request.POST.get('serialNumber')
        try:
            board = Board.objects.get(serial_number=serial_number)
            test_records = TestRecord.objects.filter(board=board, cp_nums=board.cp_nums)
            board_data = {
                'project_name': board.project_name,
                'project_config': board.project_config,
                'subproject_name': board.subproject_name,
                'serial_number': board.serial_number,
                'configuration': board.configuration,
                'board_number': board.board_number,
                'test_item_name': board.test_item_name,
                'cp_nums': board.cp_nums,
                'product_code': board.product_code,
                'APN': board.APN,
                'HHPN': board.HHPN,
                'first_GS_sn': board.first_GS_sn,
                'second_GS_sn': board.second_GS_sn,
                'env_finished_flag': board.env_finished_flag,
                'status': board.status,
            }
            test_records_data = list(test_records.values('pk', 'station_type', 'start_time', 'stop_time', 'result'))
            return JsonResponse({'board': board_data, 'test_records': test_records_data})
        except Board.DoesNotExist:
            return JsonResponse({'error': 'Serial number not found.'}, status=404)
    return JsonResponse({'error': 'Invalid request method.'}, status=400)

@csrf_exempt
def create_error_record(request):
    if request.method == 'POST':
        serial_number = request.POST.get('serial_number')
        radar = request.POST.get('radar')
        cp_nums = request.POST.get('cp_nums')
        fail_message = request.POST.get('fail_message')
        remark = request.POST.get('remark')
        test_record_id = request.POST.get('test_record')
        board_status = request.POST.get('board_status')
        print(board_status)
        fail_picture = request.FILES.get('fail_picture')

        try:
            board = Board.objects.get(serial_number=serial_number)
            if ErrorRecord.objects.filter(board=board, status='ongoing').exists():
                return JsonResponse({'error': 'There is already an ongoing ErrorRecord for this board.'}, status=400)
            test_record = TestRecord.objects.get(id=test_record_id)
            error_record = ErrorRecord.objects.create(
                board=board,
                radar=radar,
                cp_nums=cp_nums,
                status='ongoing',  # Always set to 'ongoing'
                fail_message=fail_message,
                remark=remark,
                fail_picture=fail_picture
            )
            error_record.test_record.add(test_record)
            error_record.save()
            
            board.status = board_status
            board.save()

            return JsonResponse({'success': 'Error_record created successfully.'})
        except Board.DoesNotExist:
            return JsonResponse({'error': 'Board not found.'}, status=404)
        except TestRecord.DoesNotExist:
            return JsonResponse({'error': 'Test record not found.'}, status=404)
    return JsonResponse({'error': 'Invalid request method.'}, status=400)

@csrf_exempt
def get_error_record(request, error_record_pk):
    if request.method == 'GET':
        try:
            error_record = ErrorRecord.objects.get(pk=error_record_pk)
            error_record_data = {
                'serial_number': error_record.pk,
                'radar': error_record.radar,
                'cp_nums': error_record.cp_nums,
                'fail_message': error_record.fail_message,
                'remark': error_record.remark,
                'fail_picture_url': error_record.fail_picture.url if error_record.fail_picture else None
            }
            test_records = error_record.test_record.all()
            test_records_data = list(test_records.values('pk', 'station_type', 'start_time', 'stop_time', 'result'))
            print(f"test_records_data:{test_records_data}")
            board_status = error_record.board.status
            return JsonResponse({'error_record': error_record_data, 'test_records': test_records_data, 'board_status': board_status})
        except ErrorRecord.DoesNotExist:
            return JsonResponse({'error': 'Error record not found.'}, status=404)
    return JsonResponse({'error': 'Invalid request method.'}, status=400)

@csrf_exempt
def update_error_record(request, error_record_pk):
    if request.method == 'POST':
        radar = request.POST.get('radar')
        cp_nums = request.POST.get('cp_nums')
        fail_message = request.POST.get('fail_message')
        remark = request.POST.get('remark')
        board_status = request.POST.get('board_status')
        fail_picture = request.FILES.get('fail_picture')

        try:
            error_record = ErrorRecord.objects.get(pk=error_record_pk)

            error_record.radar = radar
            error_record.cp_nums = cp_nums
            error_record.fail_message = fail_message
            error_record.remark = remark

            if fail_picture:
                # 删除旧的 fail_picture 文件
                if error_record.fail_picture:
                    if os.path.isfile(error_record.fail_picture.path):
                        os.remove(error_record.fail_picture.path)
                # 更新为新的 fail_picture 文件
                error_record.fail_picture = fail_picture

            error_record.save()

            board = error_record.board
            if board_status and board_status != "null":
                board.status = board_status
                board.save()

            for test_record in error_record.test_record.all():
                test_record.result = request.POST.get(f'test_record_result_{test_record.pk}')
                test_record.save()

            return JsonResponse({'success': 'Error record updated successfully.'})
        except ErrorRecord.DoesNotExist:
            return JsonResponse({'error': 'Error record not found.'}, status=404)
    return JsonResponse({'error': 'Invalid request method.'}, status=400)

@csrf_exempt    
def download_fail_picture(request, record_id):
    print("download_fail_picture ing~~~~~~")
    try:
        error_record = get_object_or_404(ErrorRecord, pk=record_id)
        fail_picture_path = error_record.fail_picture.path
        if os.path.exists(fail_picture_path):
            with open(fail_picture_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/octet-stream')
                response['Content-Disposition'] = f'attachment; filename={os.path.basename(fail_picture_path)}'
                return response
        return JsonResponse({'error': 'File not found.'}, status=404)
    except ErrorRecord.DoesNotExist:
        return JsonResponse({'error': 'Error record not found.'}, status=404)
    
@csrf_exempt
def get_overtime_boards(request):
    if request.method == 'GET':
        results = []
        boards = Board.objects.filter(env_finished_flag=True, if_overdue=True)
        for board in boards:
            test_record = TestRecord.objects.filter(board=board, cp_nums=board.cp_nums, station_type='checkin').get()
            time_difference = datetime.now() - test_record.start_time
            time_difference_days = time_difference.total_seconds() / (3600 * 24)
            results.append({
                "board": {
                    'project_name': board.project_name,
                    'serial_number': board.serial_number,
                    'board_number': board.board_number,
                    'env_finished_flag': board.env_finished_flag,
                    'if_overdue': board.if_overdue
                },
                "checkin_time": test_record.start_time.strftime('%Y-%m-%d %H:%M'),
                "time_difference_days": round(time_difference_days, 2)
            })
        return JsonResponse({'results': results})
    return JsonResponse({'error': 'Invalid request method.'}, status=400)

@csrf_exempt
def filter_boards(request):
    context = {}
    project_name_set = Board.objects.values_list('project_name', flat=True).distinct()
    build_name_set = Board.objects.values_list('subproject_name', flat=True).distinct()
    result_set = TestRecord.objects.values_list('result', flat=True).distinct()
    product_code_set = Board.objects.values_list('product_code', flat=True).distinct()
    site_set = Board.objects.values_list('site', flat=True).distinct()
    status_set = Board.objects.values_list('status', flat=True).distinct()
    cp_number_set = Board.objects.values_list('cp_nums', flat=True).distinct()
    project_config_set = Board.objects.values_list('project_config',flat=True).distinct()
    test_item_name_set = Board.objects.values_list('test_item_name', flat=True).distinct()
    
    context["test_item_name_set"] = test_item_name_set
    context["project_config_set"] = project_config_set
    context["cp_number_set"] = cp_number_set
    context["status_set"] = status_set
    context["site_set"] = site_set
    context["product_code_set"] = product_code_set
    context["project_name_set"] = project_name_set
    context["build_name_set"] = build_name_set
    context["result_set"] = result_set

    return render(request, "board/batch_search.html", context)

@csrf_exempt
def filter_search_boards_ajax(request):
    site = request.GET.get('site', '')
    dateTimeRange = request.GET.get('dateTimeRange', '')
    product_code = request.GET.get('product_code', '')
    project_name = request.GET.get('project_name', '')
    build_name = request.GET.get('subproject_name', '')
    result = request.GET.get('result', '')
    project_config = request.GET.get('project_config', '')
    test_item_name = request.GET.get('test_item', '')
    cp_nums = request.GET.get('cp', '')
    status = request.GET.get('status', '')
    boards = Board.objects.all()

    if site:
        boards = boards.filter(site=site)
    if dateTimeRange:
        start_time, end_time = dateTimeRange.split(' to ')
        start_time = datetime.strptime(start_time, '%Y-%m-%d')
        end_time = datetime.strptime(end_time, '%Y-%m-%d')
        test_records = TestRecord.objects.filter(start_time__range=(start_time, end_time))
        boards = boards.filter(testrecord__in=test_records).distinct()

    if product_code:
        boards = boards.filter(product_code=product_code)
    if project_name:
        boards = boards.filter(project_name=project_name)
    if build_name:
        boards = boards.filter(subproject_name=build_name)
    if result:
        boards = boards.filter(testrecord__result=result)
    if project_config:
        boards = boards.filter(project_config=project_config)
    if test_item_name:
        boards = boards.filter(test_item_name=test_item_name)
    if cp_nums:
        boards = boards.filter(cp_nums=cp_nums)
    if status:
        boards = boards.filter(status=status)

    board_sn_list = list(boards.values_list('serial_number', flat=True))
    paginator = Paginator(boards, 17)  # 每页显示17条记录
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    response = []
    for board in page_obj:
        checkin_record = board.testrecord_set.filter(station_type='checkin', cp_nums=board.cp_nums).first()
        checkout_record = board.testrecord_set.filter(station_type='checkout', cp_nums=board.cp_nums).first()
        error_record = board.errorrecord_set.first()

        response.append({
            'serial_number': board.serial_number,
            'project_config': board.project_config,
            'test_item_name': board.test_item_name,
            'product_code': board.product_code,
            'subproject_name': board.subproject_name,
            'site': board.site,
            'board_number': board.board_number,
            'configuration': board.configuration,
            'APN': board.APN,
            'HHPN': board.HHPN,
            'first_GS_sn': board.first_GS_sn,
            'second_GS_sn': board.second_GS_sn,
            'checkin_time': checkin_record.start_time.strftime('%Y/%m/%d %H:%M:%S') if checkin_record else '',
            'checkout_time': checkout_record.start_time.strftime('%Y/%m/%d %H:%M:%S') if checkout_record else '',
            'status': board.status,
            'radar': error_record.radar if error_record else '',
            'remark': error_record.remark if error_record else '',
        })

    return JsonResponse({
        'results': response,
        'total_pages': paginator.num_pages,
        'current_page': page_obj.number,
        'board_sn_list': board_sn_list
    })

@csrf_exempt
def track_board(request):
    context={}
    return render(request, 'board/single_board_track.html', context)

@csrf_exempt
def track_board_ajax(request, serial_number):
    try:
        board = Board.objects.get(serial_number=serial_number)
        test_records = TestRecord.objects.filter(board=board).order_by('-start_time')

        # 分页
        paginator = Paginator(test_records, 12)  # 每页显示12条记录
        page_number = request.GET.get('page', 1)  # 获取当前页码，默认为第一页
        page_obj = paginator.get_page(page_number)

        board_data = {
            'serial_number': board.serial_number,
            'subproject_name': board.subproject_name,
            'test_item_name': board.test_item_name,
            'product_code': board.product_code,
            'site': board.site,
            'board_number': board.board_number,
            'configuration': board.configuration,
            'APN': board.APN,
            'first_GS_sn': board.first_GS_sn,
            'second_GS_sn': board.second_GS_sn,
            'status': board.status,
            'remark': "remark",
        }

        test_records_data = []
        for record in page_obj:
            test_records_data.append({
                'station_type': record.station_type,
                'start_time': record.start_time.strftime('%Y/%m/%d %H:%M:%S'),
                'stop_time': record.stop_time.strftime('%Y/%m/%d %H:%M:%S'),
                'cp_nums': record.cp_nums,
                'result': record.result,
                'operator': record.operator,
                'remark': record.remark,
            })

        return JsonResponse({
            'board': board_data,
            'test_records': test_records_data,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number,
        })

    except Board.DoesNotExist:
        return JsonResponse({'error': '没找到该SN的板子'}, status=404)
    
@csrf_exempt
def export_report(request):
    site = request.GET.get('site', '')
    dateTimeRange = request.GET.get('dateTimeRange', '')
    product_code = request.GET.get('product_code', '')
    project_name = request.GET.get('project_name', '')
    subproject_name = request.GET.get('subproject_name', '')
    result = request.GET.get('result', '')

    boards = Board.objects.all()

    if site:
        boards = boards.filter(site=site)
    if dateTimeRange:
        start_time, end_time = dateTimeRange.split(' to ')
        start_time = datetime.strptime(start_time, '%Y-%m-%d')
        end_time = datetime.strptime(end_time, '%Y-%m-%d')
        test_records = TestRecord.objects.filter(start_time__range=(start_time, end_time))
        boards = boards.filter(testrecord__in=test_records).distinct()
    if product_code:
        boards = boards.filter(product_code=product_code)
    if project_name:
        boards = boards.filter(project_name=project_name)
    if subproject_name:
        boards = boards.filter(subproject_name=subproject_name)
    if result:
        boards = boards.filter(testrecord__result=result)

    # Create a workbook and add a worksheet.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Add headers
    headers = [
        'Test Catagory', 'Test', 'Chamber', 'Product Code', 'Build Stage', 'Site',
        'Unit Number', 'MLB Config', 'APN', 'Nand_SN0', 'Nand_SN1',
        'Check-In Time', 'Check-Out Time', 'Board Status', 'Radar', 'Remark'
    ]
    ws.append(headers)

    # Add data
    for board in boards:
        checkin_record = board.testrecord_set.filter(station_type='checkin', cp_nums=board.cp_nums).first()
        checkout_record = board.testrecord_set.filter(station_type='checkout', cp_nums=board.cp_nums).first()
        error_records = board.errorrecord_set.all()

        radar = '\n'.join([record.radar for record in error_records if record.radar])
        remark = '\n'.join([record.remark for record in error_records if record.remark])

        row = [
            board.project_config,
            board.test_item_name,
            '/',
            board.product_code,
            board.subproject_name,
            board.site,
            board.board_number,
            board.configuration,
            board.APN,
            board.first_GS_sn,
            board.second_GS_sn,
            checkin_record.start_time.strftime('%Y/%m/%d %H:%M:%S') if checkin_record else '',
            checkout_record.start_time.strftime('%Y/%m/%d %H:%M:%S') if checkout_record else '',
            board.status,
            radar,
            remark,
        ]
        ws.append(row)

    # Set column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Create a response
    current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'report_{current_time}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response

@csrf_exempt
def confirm_archive_boards(request):
    if request.method == 'POST':
        selected_boards = json.loads(request.POST.get('selected_boards', '[]'))
        userId = request.POST.get('userId')
        print("userId:"+userId)
        # 在此处理选中的板子信息，例如更新数据库中的字段以标记为已存档
        for serial_number in selected_boards:
            print(serial_number)
            board = Board.objects.get(serial_number=serial_number)
            record = TestRecord(board=board, station_type='archived', result='pass', operator=userId, start_time = datetime.now(), stop_time = datetime.now(), cp_nums = board.cp_nums).save()
            record.save()
            board.status = 'archived'
            board.save()
        return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'failed'}, status=400)

@csrf_exempt
def download_archive_boards(request):
    if request.method == 'POST':
        selected_boards = json.loads(request.POST.get('selected_boards', '[]'))
        boards = Board.objects.filter(serial_number__in=selected_boards)

        # 创建 Excel 文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Boards Info'
        
        # 添加表头
        headers = ['Serial Number', 'Project Config', 'Test Item Name', 'Product Code', 'Subproject Name', 'Site']
        ws.append(headers)

        # 添加数据
        for board in boards:
            ws.append([
                board.serial_number,
                board.project_config,
                board.test_item_name,
                board.product_code,
                board.subproject_name,
                board.site
            ])
        
        # Set column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=boards_info.xlsx'
        wb.save(response)
        
        return response

    return JsonResponse({'status': 'failed'}, status=400)

@csrf_exempt
def archive_boards(request):
    context = {}
    project_name_set = Board.objects.values_list('project_name', flat=True).distinct()
    build_name_set = Board.objects.values_list('subproject_name', flat=True).distinct()
    result_set = TestRecord.objects.values_list('result', flat=True).distinct()
    product_code_set = Board.objects.values_list('product_code', flat=True).distinct()
    site_set = Board.objects.values_list('site', flat=True).distinct()
    status_set = Board.objects.values_list('status', flat=True).distinct()
    cp_number_set = Board.objects.values_list('cp_nums', flat=True).distinct()
    project_config_set = Board.objects.values_list('project_config',flat=True).distinct()
    test_item_name_set = Board.objects.values_list('test_item_name', flat=True).distinct()
    
    context["test_item_name_set"] = test_item_name_set
    context["project_config_set"] = project_config_set
    context["cp_number_set"] = cp_number_set
    context["status_set"] = status_set
    context["site_set"] = site_set
    context["product_code_set"] = product_code_set
    context["project_name_set"] = project_name_set
    context["build_name_set"] = build_name_set
    context["result_set"] = result_set

    return render(request, "board/archive_board.html", context)

@csrf_exempt
def archive_boards_ajax(request):
    site = request.GET.get('site', '')
    dateTimeRange = request.GET.get('dateTimeRange', '')
    product_code = request.GET.get('product_code', '')
    project_name = request.GET.get('project_name', '')
    build_name = request.GET.get('subproject_name', '')
    result = request.GET.get('result', '')
    project_config = request.GET.get('project_config', '')
    test_item_name = request.GET.get('test_item', '')
    cp_nums = request.GET.get('cp', '')
    status = request.GET.get('status', '')
    
    print("archive board____________________")
    
    boards = Board.objects.all()

    if site:
        boards = boards.filter(site=site)
    if dateTimeRange:
        start_time, end_time = dateTimeRange.split(' to ')
        start_time = datetime.strptime(start_time, '%Y-%m-%d')
        end_time = datetime.strptime(end_time, '%Y-%m-%d')
        test_records = TestRecord.objects.filter(start_time__range=(start_time, end_time))
        boards = boards.filter(testrecord__in=test_records).distinct()

    if product_code:
        boards = boards.filter(product_code=product_code)
    if project_name:
        boards = boards.filter(project_name=project_name)
    if build_name:
        boards = boards.filter(subproject_name=build_name)
    if result:
        boards = boards.filter(testrecord__result=result)
    if project_config:
        boards = boards.filter(project_config=project_config)
    if test_item_name:
        boards = boards.filter(test_item_name=test_item_name)
    if cp_nums:
        boards = boards.filter(cp_nums=cp_nums)
    if status:
        boards = boards.filter(status=status)

    board_sn_list = list(boards.values_list('serial_number', flat=True))
    paginator = Paginator(boards, 17)  # 每页显示17条记录
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    response = []
    for board in page_obj:
        checkin_record = board.testrecord_set.filter(station_type='checkin', cp_nums=board.cp_nums).first()
        checkout_record = board.testrecord_set.filter(station_type='checkout', cp_nums=board.cp_nums).first()
        error_record = board.errorrecord_set.first()

        response.append({
            'serial_number': board.serial_number,
            'project_config': board.project_config,
            'test_item_name': board.test_item_name,
            'product_code': board.product_code,
            'subproject_name': board.subproject_name,
            'site': board.site,
            'board_number': board.board_number,
            'configuration': board.configuration,
            'APN': board.APN,
            'HHPN': board.HHPN,
            'first_GS_sn': board.first_GS_sn,
            'second_GS_sn': board.second_GS_sn,
            'checkin_time': checkin_record.start_time.strftime('%Y/%m/%d %H:%M:%S') if checkin_record else '',
            'checkout_time': checkout_record.start_time.strftime('%Y/%m/%d %H:%M:%S') if checkout_record else '',
            'status': board.status,
            'radar': error_record.radar if error_record else '',
            'remark': error_record.remark if error_record else '',
        })

    return JsonResponse({
        'results': response,
        'total_pages': paginator.num_pages,
        'current_page': page_obj.number,
        'board_sn_list': board_sn_list
    })